import argparse
import json
import logging
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib import error, parse, request

from openpyxl import load_workbook


LOG_FORMAT = "%(asctime)s | %(levelname)s | %(message)s"
logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)
logger = logging.getLogger("import_sinapi_pr_sem_desoneracao")


DEFAULT_REFERENCIA_FILE = r"orcamentista/templates/SINAPI_Referência_2026_03.xlsx"
DEFAULT_MAO_DE_OBRA_FILE = r"orcamentista/templates/SINAPI_mao_de_obra_2026_03.xlsx"
DEFAULT_MANUTENCOES_FILE = r"orcamentista/templates/SINAPI_Manutenções_2026_03.xlsx"
DEFAULT_METADATA_OUT = r"orcamentista/sinapi/import_metadata_pr_sem_desoneracao.json"
DEFAULT_LOG_DIR = r"orcamentista/sinapi/logs"

CODE_FORMULA_RE = re.compile(r",(\d+)\)\s*$")


def resolve_default_log_file() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(Path(DEFAULT_LOG_DIR) / f"import_sinapi_{timestamp}.log")


def configure_file_logging(log_file: str) -> str:
    log_path = Path(log_file).expanduser()
    log_path.parent.mkdir(parents=True, exist_ok=True)

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(LOG_FORMAT))
    logging.getLogger().addHandler(file_handler)

    return str(log_path)


def strip_optional_quotes(value: str) -> str:
    if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
        return value[1:-1]
    return value


def discover_env_file(explicit_path: Optional[str] = None) -> Optional[Path]:
    if explicit_path:
        env_path = Path(explicit_path).expanduser()
        return env_path if env_path.exists() else None

    script_path = Path(__file__).resolve()
    candidates = [
        Path.cwd() / ".env",
        script_path.parent / ".env",
        script_path.parent.parent / ".env",
        script_path.parent.parent.parent / ".env",
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    return None


def load_env_file(explicit_path: Optional[str] = None) -> Optional[str]:
    env_path = discover_env_file(explicit_path)
    if env_path is None:
        return None

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        if line.startswith("export "):
            line = line[7:].strip()

        if "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        if not key:
            continue

        os.environ.setdefault(key, strip_optional_quotes(value.strip()))

    return str(env_path)


class SupabaseImportClient:
    def __init__(self, client: Any):
        self.client = client

    def upsert(self, table_name: str, records: List[dict], on_conflict: str) -> None:
        response = (
            self.client.table(table_name)
            .upsert(records, on_conflict=on_conflict)
            .execute()
        )
        error_response = getattr(response, "error", None)
        if error_response:
            raise RuntimeError(f"Erro retornado pelo Supabase: {error_response}")


class PostgrestImportClient:
    def __init__(self, url: str, key: str):
        self.base_url = url.rstrip("/")
        self.key = key

    def upsert(self, table_name: str, records: List[dict], on_conflict: str) -> None:
        endpoint = f"{self.base_url}/rest/v1/{table_name}"
        query = parse.urlencode({"on_conflict": on_conflict})
        payload = json.dumps(records, ensure_ascii=False).encode("utf-8")

        req = request.Request(
            url=f"{endpoint}?{query}",
            data=payload,
            method="POST",
            headers={
                "apikey": self.key,
                "Authorization": f"Bearer {self.key}",
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
        )

        try:
            with request.urlopen(req, timeout=120):
                return
        except error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(
                f"Erro HTTP ao inserir lote no Supabase ({exc.code}): {body}"
            ) from exc
        except error.URLError as exc:
            raise RuntimeError(f"Falha de rede ao acessar o Supabase: {exc}") from exc


def clean_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def parse_competencia(value: object) -> str:
    text = clean_text(value)
    if not text:
        raise ValueError("Competencia nao encontrada")

    if re.match(r"^\d{2}/\d{4}$", text):
        month, year = text.split("/")
        return f"{year}-{month}-01"

    if isinstance(value, datetime):
        return value.date().replace(day=1).isoformat()

    raise ValueError(f"Competencia em formato inesperado: {value}")


def get_supabase_client() -> Any:
    url = os.getenv("SUPABASE_URL") or os.getenv("VITE_SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        extra_hint = ""
        if os.getenv("VITE_SUPABASE_ANON_KEY") and not key:
            extra_hint = (
                " Foi encontrada apenas VITE_SUPABASE_ANON_KEY. "
                "Para importacao use SUPABASE_KEY ou SUPABASE_SERVICE_ROLE_KEY."
            )
        raise RuntimeError(
            "SUPABASE_URL (ou VITE_SUPABASE_URL) e uma chave de servico sao obrigatorios."
            + extra_hint
        )

    try:
        from supabase import create_client

        logger.info("Cliente Supabase inicializado via supabase-py.")
        return SupabaseImportClient(create_client(url, key))
    except Exception as exc:
        logger.warning(
            "supabase-py indisponivel neste ambiente (%s). Usando fallback HTTP via PostgREST.",
            exc,
        )
        return PostgrestImportClient(url, key)


def find_value_column(ws, header_row: int, label: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(header_row, col).value == label:
            return col
    raise ValueError(f"Coluna '{label}' nao encontrada na aba {ws.title}")


def extract_code_from_formula(value: object) -> Optional[str]:
    if isinstance(value, str):
        match = CODE_FORMULA_RE.search(value)
        if match:
            return match.group(1)
    return None


def load_csd_pr(reference_file: str) -> Tuple[Dict[str, dict], str]:
    wb_formula = load_workbook(reference_file, read_only=True, data_only=False)
    wb_values = load_workbook(reference_file, read_only=True, data_only=True)

    ws_formula = wb_formula["CSD"]
    ws_values = wb_values["CSD"]

    competencia = parse_competencia(ws_values["B3"].value)
    pr_col = find_value_column(ws_formula, 9, "PR")
    pr_as_col = pr_col + 1

    records: Dict[str, dict] = {}
    for row_formula, row_values in zip(
        ws_formula.iter_rows(min_row=11, values_only=True),
        ws_values.iter_rows(min_row=11, values_only=True),
    ):
        codigo = extract_code_from_formula(row_formula[1] if len(row_formula) > 1 else None)
        descricao = clean_text(row_formula[2] if len(row_formula) > 2 else None)
        unidade = clean_text(row_formula[3] if len(row_formula) > 3 else None)
        categoria = clean_text(row_formula[0] if len(row_formula) > 0 else None)

        if not codigo or not descricao or not unidade:
            continue

        valor_unitario = row_values[pr_col - 1] if len(row_values) >= pr_col else None
        percentual_as = row_values[pr_as_col - 1] if len(row_values) >= pr_as_col else None

        records[codigo] = {
            "codigo": codigo,
            "descricao": descricao,
            "unidade": unidade,
            "categoria": categoria,
            "uf": "PR",
            "regime_desoneracao": "SEM_DESONERACAO",
            "competencia": competencia,
            "valor_unitario": float(valor_unitario) if valor_unitario is not None else None,
            "percentual_atribuido_sp": float(percentual_as) if percentual_as is not None else None,
            "percentual_mao_de_obra": None,
            "situacao": None,
            "composicao": [],
            "manutencoes": [],
            "origem": "SINAPI",
            "ativo": True,
        }

    return records, competencia


def load_analitico(reference_file: str) -> Dict[str, dict]:
    wb = load_workbook(reference_file, read_only=True, data_only=True)
    ws = wb["Analítico"]

    result: Dict[str, dict] = {}
    for row in ws.iter_rows(min_row=11, values_only=True):
        codigo_comp = row[1]
        tipo_item = row[2]
        codigo_item = row[3]
        descricao = clean_text(row[4])
        unidade = clean_text(row[5])
        coeficiente = row[6]
        situacao = clean_text(row[7])
        categoria = clean_text(row[0])

        if codigo_comp is None or not descricao:
            continue

        codigo_comp = str(int(codigo_comp))
        bucket = result.setdefault(
            codigo_comp,
            {"situacao": None, "categoria": categoria, "itens": []},
        )

        if tipo_item is None:
            bucket["situacao"] = situacao
            if categoria and not bucket.get("categoria"):
                bucket["categoria"] = categoria
            continue

        bucket["itens"].append(
            {
                "tipo_item": clean_text(tipo_item),
                "codigo_item": str(int(codigo_item)) if codigo_item is not None else None,
                "descricao": descricao,
                "unidade": unidade,
                "coeficiente": float(coeficiente) if coeficiente is not None else None,
                "situacao": situacao,
            }
        )

    return result


def load_mao_de_obra_pr(mao_file: str) -> Dict[str, float]:
    wb = load_workbook(mao_file, read_only=True, data_only=True)
    ws = wb["SEM Desoneração"]
    pr_col = find_value_column(ws, 6, "PR")

    result: Dict[str, float] = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        codigo = row[1]
        descricao = row[2]
        percentual = row[pr_col - 1]
        if codigo is None or descricao in (None, ""):
            continue
        if percentual is None:
            continue
        result[str(int(codigo))] = float(percentual)

    return result


def load_manutencoes(manutencoes_file: str, competencia: str) -> Dict[str, List[dict]]:
    wb = load_workbook(manutencoes_file, read_only=True, data_only=True)
    ws = wb["Manutenções"]

    result: Dict[str, List[dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=7, values_only=True):
        referencia, tipo, codigo, descricao, manutencao = row[:5]
        if not tipo or not codigo:
            continue
        if "COMPOS" not in str(tipo).upper():
            continue

        result[str(int(codigo))].append(
            {
                "competencia_evento": referencia.date().isoformat() if isinstance(referencia, datetime) else competencia,
                "tipo_registro": clean_text(tipo),
                "descricao_evento": clean_text(descricao),
                "manutencao": clean_text(manutencao),
            }
        )

    return result


def build_records(
    referencia_file: str,
    mao_file: str,
    manutencoes_file: str,
) -> Tuple[List[dict], Dict[str, object]]:
    base_records, competencia = load_csd_pr(referencia_file)
    analitico = load_analitico(referencia_file)
    mao = load_mao_de_obra_pr(mao_file)
    manut = load_manutencoes(manutencoes_file, competencia)

    for codigo, record in base_records.items():
        if codigo in analitico:
            record["situacao"] = analitico[codigo]["situacao"]
            record["composicao"] = analitico[codigo]["itens"]
            if not record["categoria"] and analitico[codigo].get("categoria"):
                record["categoria"] = analitico[codigo]["categoria"]

        if codigo in mao:
            record["percentual_mao_de_obra"] = mao[codigo]

        if codigo in manut:
            record["manutencoes"] = manut[codigo]

    metadata = {
        "uf": "PR",
        "regime_desoneracao": "SEM_DESONERACAO",
        "competencia": competencia,
        "total_composicoes_csd": len(base_records),
        "com_analitico": sum(1 for r in base_records.values() if r["composicao"]),
        "com_percentual_mao_de_obra": sum(
            1 for r in base_records.values() if r["percentual_mao_de_obra"] is not None
        ),
        "com_manutencoes": sum(1 for r in base_records.values() if r["manutencoes"]),
        "arquivos": {
            "referencia": referencia_file,
            "mao_de_obra": mao_file,
            "manutencoes": manutencoes_file,
        },
    }

    return list(base_records.values()), metadata


def batch_insert(supabase: Any, records: List[dict], batch_size: int = 200) -> int:
    total = len(records)
    inserted = 0
    for start in range(0, total, batch_size):
        batch = records[start : start + batch_size]
        supabase.upsert(
            "sinapi_composicoes",
            batch,
            on_conflict="codigo,uf,regime_desoneracao,competencia",
        )

        inserted += len(batch)
        logger.info("Lote inserido: %s/%s", inserted, total)

    return inserted


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Importa SINAPI PR sem desoneracao para a tabela sinapi_composicoes"
    )
    parser.add_argument("--referencia-file", default=DEFAULT_REFERENCIA_FILE)
    parser.add_argument("--mao-file", default=DEFAULT_MAO_DE_OBRA_FILE)
    parser.add_argument("--manutencoes-file", default=DEFAULT_MANUTENCOES_FILE)
    parser.add_argument("--metadata-out", default=DEFAULT_METADATA_OUT)
    parser.add_argument("--batch-size", type=int, default=200)
    parser.add_argument("--env-file")
    parser.add_argument("--log-file")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    start = datetime.now()
    log_file = configure_file_logging(args.log_file or resolve_default_log_file())
    env_file = load_env_file(args.env_file)

    if env_file:
        logger.info("Arquivo .env carregado de: %s", env_file)
    else:
        logger.info("Nenhum arquivo .env encontrado. Variaveis de ambiente do sistema serao usadas.")

    records, metadata = build_records(
        referencia_file=args.referencia_file,
        mao_file=args.mao_file,
        manutencoes_file=args.manutencoes_file,
    )
    metadata["log_file"] = log_file
    metadata["env_file"] = env_file

    with open(args.metadata_out, "w", encoding="utf-8") as fp:
        json.dump(metadata, fp, ensure_ascii=False, indent=2)

    logger.info("Total de composicoes preparadas: %s", len(records))
    logger.info("Metadata salva em: %s", args.metadata_out)
    logger.info("Log de execucao salvo em: %s", log_file)

    if args.dry_run:
        logger.info("Dry run ativo. Nenhum dado enviado ao Supabase.")
        return

    supabase = get_supabase_client()
    inserted = batch_insert(supabase, records, args.batch_size)
    logger.info("Importacao concluida com %s registros enviados ao Supabase.", inserted)

    elapsed = datetime.now() - start
    logger.info("Tempo total de execucao: %s", elapsed)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        logger.exception("Falha na importacao: %s", exc)
        sys.exit(1)
